from single_engine_predict_iodie import *
from MappingGen import *
from mesh_hetero import *

import os
import yaml
import openpyxl
import argparse

cur_dir = os.path.dirname(os.path.abspath(__file__))

def yamlLoad(file):
    with open(file, encoding='utf-8') as rstream:
        data = yaml.load(rstream, yaml.SafeLoader)
    return data

# SampleGen, iterTimes, memory_param, NoC_param, all_sim_node_num , if_multicast, excel_filename, workload_name, flag = "ours"
def Explore(arch, dimension_tag, parallel_type, workload, chiplet_size, temporal_level, iterTimes, NoC_param, if_multicast, objective):
    # setting some parameters
    HW_param = {"Chiplet": chiplet_size, "PE": arch["PE"], "MAC": arch["MAC"]}
    memory_param = {}
    memory_param["OL1"] = arch["L1"]['o']
    memory_param["AL1"] = arch["L1"]['i']
    memory_param["WL1"] = arch["L1"]['w']
    memory_param["OL2"] = arch["L2"]['o']
    memory_param["AL2"] = arch["L2"]['i']
    memory_param["WL2"] = arch["L2"]['w']
    
    perIters = math.ceil(iterTimes / len(parallel_type["Chiplet"]) / len(parallel_type["PE"]))
    
    # records
    record = {}
    record["title"] =   [ \
                            "for_list", "comm_patterns", "parallel_dim_list", "partition_list", "edp", "energy", "latency", "degrade_ratio", "degrade_ratio_dict", \
                            "pkt_needed", "neu_needed", "compuation_cycles", "runtime_list", "cp_list", "utilization_ratio_list", "energy_dram_list", \
                            "energy_L2_list", "energy_L1_list", "energy_die2die", "energy_MAC", "energy_psum_list", "delay_psum", "worstlinks" \
                        ]
    nn_best_fitness     = {}
    nn_best_performance = {"edp":{}, "energy":{}, "latency":{}}
    nn_best_degrade_ratio = {}
    nn_best_sample      = {}
    nn_best_sp_type     = {}
    nn_best_pkt_needed  = {}
    nn_best_neu_needed  = {}
    # MappingGenerator Initialize
    MG = MappingGenerator(temporal_level, HW_param, dimension_tag)
    for layer_name, layer in workload.items():
        record[layer_name] = []
        MG.set_layer(layer)
        
        best_fitness = None
        best_performance = None
        best_degrade_ratio = None
        best_sample = None
        best_sp_type = None
        best_pkt_needed = None
        best_neu_needed = None
        for chiplet_sp_type in parallel_type["Chiplet"]:
            for PE_sp_type in parallel_type["PE"]:
                # 设置并行度
                MG.set_spatial_parallel(Chiplet_type=chiplet_sp_type, PE_type=PE_sp_type)
                MG.set_temporal_dimensions()
                # 迭代优化
                for i in range(perIters):
                    # --- 生成新样本
                    for_list, comm_patterns, parallel_dim_list, partition_list = MG.fresh()
                    # --- 样本性能评估
                    latency, degrade_ratio, degrade_ratio_dict, pkt_needed, neu_needed, compuation_cycles, runtime_list, cp_list, utilization_ratio_list, \
                    energy_dram_list, energy_L2_list,energy_L1_list, energy_die2die, energy_MAC, energy_psum_list, delay_psum, worstlinks \
                    = calFitness(for_list, comm_patterns, parallel_dim_list, partition_list, layer, HW_param, memory_param, NoC_param, if_multicast)
                    # --- 比较
                    e_mem = sum(energy_dram_list)+sum(energy_L2_list)+sum(energy_L1_list)
                    energy = e_mem + energy_die2die+energy_MAC + energy_psum_list[2]
                    edp = latency * energy  /(PE_freq * freq_1G) # pJ*s
                    
                    fitness_list = {"edp": edp, "energy": energy, "latency": latency}
                    fitness = fitness_list[objective]
                    if best_fitness == None or fitness < best_fitness:
                        best_fitness        = fitness
                        best_performance    = fitness_list
                        best_pkt_needed     = copy.deepcopy(pkt_needed)
                        best_neu_needed     = copy.deepcopy(neu_needed)
                        best_sp_type        = "c_{}_pe_{}".format(chiplet_sp_type, PE_sp_type)
                        best_degrade_ratio  = copy.deepcopy(degrade_ratio_dict)
                        best_sample                         = {}
                        best_sample["for_list"]             = copy.deepcopy(for_list)
                        best_sample["comm_patterns"]        = copy.deepcopy(comm_patterns)
                        best_sample["parallel_dim_list"]    = copy.deepcopy(parallel_dim_list)
                        best_sample["partition_list"]       = copy.deepcopy(partition_list)
                        best_sample["fitness_list"]         = copy.deepcopy(fitness_list)
                        best_sample["degrade_ratio_dict"]   = copy.deepcopy(degrade_ratio_dict)
             
                    print("fitness_min = {}, sp_type = {}, degrade_ratio = {}".format(best_fitness, best_sp_type, str(best_sample["degrade_ratio_dict"] )))
                        
                    # --- 记录
                    record[layer_name].append( \
                        [for_list, comm_patterns, parallel_dim_list, partition_list, edp, energy, latency, degrade_ratio, degrade_ratio_dict, \
                        pkt_needed, neu_needed, compuation_cycles, runtime_list, cp_list, utilization_ratio_list, energy_dram_list, energy_L2_list, \
                        energy_L1_list, energy_die2die, energy_MAC, energy_psum_list, delay_psum, worstlinks] \
                    )
        
        nn_best_fitness[layer_name]                 = copy.deepcopy(best_fitness)
        nn_best_performance["edp"][layer_name]      = best_performance["edp"]
        nn_best_performance["energy"][layer_name]   = best_performance["energy"]
        nn_best_performance["latency"][layer_name]  = best_performance["latency"]
        nn_best_degrade_ratio[layer_name]           = copy.deepcopy(best_degrade_ratio)
        nn_best_sample[layer_name]                  = copy.deepcopy(best_sample)
        nn_best_sp_type[layer_name]                 = copy.deepcopy(best_sp_type)
        nn_best_pkt_needed[layer_name]              = copy.deepcopy(best_pkt_needed)
        nn_best_neu_needed[layer_name]              = copy.deepcopy(best_neu_needed)
        
    return nn_best_fitness, nn_best_performance, nn_best_degrade_ratio, nn_best_sample, nn_best_sp_type, nn_best_pkt_needed, nn_best_neu_needed, record

def ExtractConfig(architecture, df, nn_name):
    config_dir = os.path.join(cur_dir, 'config')
    arch_file = os.path.join(config_dir, 'arch.yaml')
    dataflow_file = os.path.join(config_dir, 'dataflow.yaml')
    workload_file = os.path.join(config_dir, 'workload/{}.yaml'.format(nn_name))
    arch = yamlLoad(arch_file)
    dataflow = yamlLoad(dataflow_file)
    workload = yamlLoad(workload_file)
    layer_name_dict = workload.pop("layer_name_dict")
    dimension_tag = dataflow['{}_dimension_tag'.format(df)]
    parallel_type = dataflow['{}_parallel_type'.format(df)]
    return arch[architecture], dimension_tag, parallel_type, workload, layer_name_dict

def set_chiplet_size(total_num):
    chiplet_size = {}
    if total_num == 16:
        chiplet_size[1] = {'x': 1, 'y': 1}
        chiplet_size[2] = {'x': 2, 'y': 1}
        chiplet_size[3] = {'x': 3, 'y': 1}
        chiplet_size[4] = {'x': 2, 'y': 2}
        chiplet_size[5] = {'x': 5, 'y': 1}
        chiplet_size[6] = {'x': 2, 'y': 3}
        chiplet_size[7] = {'x': 7, 'y': 1}
        chiplet_size[8] = {'x': 4, 'y': 2}
        chiplet_size[9] = {'x': 3, 'y': 3}
        chiplet_size[10] = {'x': 2, 'y': 5}
        chiplet_size[11] = {'x': 11, 'y': 1}
        chiplet_size[12] = {'x': 3, 'y': 4}
        chiplet_size[13] = {'x': 13, 'y': 1}
        chiplet_size[14] = {'x': 2, 'y': 7}
        chiplet_size[15] = {'x': 3, 'y': 5}
        chiplet_size[16] = {'x': 4, 'y': 4}
    elif total_num == 36:
        chiplet_size[6] = {'x': 3, 'y': 2}
        chiplet_size[12] = {'x': 3, 'y': 4}
        chiplet_size[18] = {'x': 3, 'y': 6}
        chiplet_size[24] = {'x': 4, 'y': 6}
        chiplet_size[30] = {'x': 5, 'y': 6}
        chiplet_size[36] = {'x': 6, 'y': 6}
    return chiplet_size

def save_record(records, out_dir):
    titles = records.pop("title")
    for layer_name, record in records.items:
        file_name = "{}/{}.xlsx".format(out_dir, layer_name)
        workbook = openpyxl.Workbook()
        sheet = workbook.get_sheet_by_name('Sheet')
        for col, title in enumerate(titles):
            sheet.cell(1, col+1, title)
        
        for row, data in enumerate(record):
            for col, item in enumerate(data):
                sheet.cell(row+2, col+1, item)
        workbook.save(file_name)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--architecture', type=str, default="ours", help='hardware architecture type (ours, nnbaton, simba)')    # simba , nnbaton
    parser.add_argument('--nn_name', type=str, default="resnet50", help='NN model name')
    parser.add_argument('--dataflow', type=str, default="os", help='dataflow type (os)')
    parser.add_argument('--chiplet_num_max', type=int, default=None, help='max chiplet_num')
    parser.add_argument('--chiplet_num_min', type=int, default=None, help='min chiplet_num')
    parser.add_argument('--chiplet_num_gran', type=int, default=None, help='chiplet_num granurity')
    parser.add_argument('--save_all_records', type=int, default=0, help='save all record')
    parser.add_argument('--optimization_objective', type=str, default="edp", help='optimization_objective')
    parser.add_argument('--iterTimes', type=int, default=1000, help='iteration times')
    parser.add_argument('--temporal_level', type=int, default=3, help='temporal level')
    parser.add_argument('--multicast_tag', type=int, default=1, help='multicast open or not')
    opt = parser.parse_args()
    
    architecture = opt.architecture
    df = opt.dataflow
    nn_name = opt.nn_name
    
    chiplet_num_max = opt.chiplet_num_max
    chiplet_num_min = opt.chiplet_num_min
    chiplet_num_gran = opt.chiplet_num_gran
    iterTimes = opt.iterTimes
    if_multicast = opt.multicast_tag
    objective = opt.optimization_objective
    temporal_level = opt.temporal_level
    save_all_records = opt.save_all_records
    
    result_dir = os.path.join(cur_dir, "result")
    os.makedirs(result_dir, exist_ok=True)
    result_dir = os.path.join(result_dir, "{}_{}".format(architecture, nn_name))
    os.makedirs(result_dir, exist_ok=True)
    result_dir = os.path.join(result_dir, objective)
    os.makedirs(result_dir, exist_ok=True)
                
    chiplet_size_dict = set_chiplet_size(chiplet_num_max)
    arch, dimension_tag, parallel_type, workload, layer_name_dict = ExtractConfig(architecture, df, nn_name)
    
    NOC_NODE_NUM = (arch["PE"]['x']+1) * arch["PE"]['y']
    NOP_SIZE = 1
    NoC_w = arch["PE"]['x']
    NoP_w = 1
    nop_scale_ratio = 1
    NoC_param, all_SIM_NODE_NUM = construct_noc_nop_Mesh(NOC_NODE_NUM, NoC_w, NOP_SIZE, NoP_w, nop_scale_ratio)
    
    for c_num in range(chiplet_num_min, chiplet_num_max + 1, chiplet_num_gran):
        
        # 探索
        nn_best_fitness, nn_best_performance, nn_best_degrade_ratio,  nn_best_sample, nn_best_sp_type, nn_best_pkt_needed, nn_best_neu_needed, record = \
            Explore(arch, dimension_tag, parallel_type, workload, chiplet_size_dict[c_num], temporal_level, iterTimes, NoC_param, if_multicast, objective)
        
        # 结果记录
        file_1 = os.path.join(result_dir, "{}_final_result_record.txt".format(c_num))
        f = open(file_1,'w')
        print("edp: ", nn_best_performance["edp"], file=f)
        print("energy: ", nn_best_performance["energy"], file=f)
        print("latency: ", nn_best_performance["latency"], file=f)
        print("degrade_ratio: ", nn_best_degrade_ratio, file = f)
        print("layer_name_dict: ", layer_name_dict, file=f)
        f.close()
        
        file_2 = os.path.join(result_dir, "{}_pkt_neu_needed.txt".format(c_num))
        f = open(file_2,'w')
        for layer_name in nn_best_pkt_needed:
            pkt_needed = nn_best_pkt_needed[layer_name]
            neu_needed = nn_best_neu_needed[layer_name]
            print("workload: {} {}".format(nn_name, layer_name), file=f)
            print("pkt_needed: ", pkt_needed, file=f)
            print("neu_needed: ", neu_needed, file=f)
        f.close()
        
        if save_all_records == 1:
            out_dir = '{}/{}'.format(cur_dir, "sample_record")
            os.makedirs(out_dir, exist_ok=True)
            out_dir = os.path.join(out_dir, nn_name)
            os.makedirs(out_dir, exist_ok=True)
            out_dir = os.path.join(out_dir, "c_{}".format(c_num))
            os.makedirs(out_dir, exist_ok=True)
            save_record(record, out_dir)
        